import os
import sqlite3
import threading
from typing import Optional, List, Tuple, Dict, Any
from datetime import datetime
from openpyxl import load_workbook

from .config import SQLITE_DB_PATH, SQLITE_DIR
from .utils import normalize_word
from .loading import loading_state
from .logger import get_logger

logger = get_logger(__name__)

# 加载状态（只保留加载相关的，其他用 app_state）
loading_thread: Optional[threading.Thread] = None
loading_cancelled = False

# 备注：状态统一放在 core/app_state.py 中


def compute_total_rows(file_path: str) -> int:
    """计算 Excel 文件总行数"""
    try:
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)
        total = 0
        for ws in wb.worksheets:
            total += (ws.max_row or 0)
        return total
    except Exception:
        return 0


def _rebuild_sqlite_from_excel(file_path: str) -> None:
    """从 Excel 重建 SQLite 数据库"""
    global loading_cancelled
    from .app_state import app_state
    
    wb = load_workbook(filename=file_path, read_only=True, data_only=True)
    sheets = [ws.title for ws in wb.worksheets]

    os.makedirs(SQLITE_DIR, exist_ok=True)

    con = sqlite3.connect(SQLITE_DB_PATH)
    cur = con.cursor()
    try:
        cur.execute("PRAGMA journal_mode=WAL;")
        cur.execute("PRAGMA synchronous=NORMAL;")
        cur.execute("PRAGMA temp_store=MEMORY;")

        cur.execute("DROP TABLE IF EXISTS entries;")
        cur.execute(
            """
            CREATE TABLE entries (
              id INTEGER PRIMARY KEY,
              word_norm TEXT NOT NULL,
              word TEXT,
              phonetic TEXT,
              meaning TEXT,
              sheet TEXT,
              row_index INTEGER
            );
            """
        )

        con.execute("BEGIN;")

        SAMPLE_STEP = 10
        LATEST_LIMIT = 40
        loading_state.clear_error()

        insert_sql = (
            "INSERT INTO entries (word_norm, word, phonetic, meaning, sheet, row_index) VALUES (?, ?, ?, ?, ?, ?)"
        )

        for sheet in sheets:
            if loading_cancelled:
                con.rollback()
                raise RuntimeError("loading cancelled")
                
            loading_state.set_current_sheet(sheet)
            ws = wb[sheet]
            max_cols = ws.max_column or 1
            word_col_idx = 1 if max_cols > 1 else 0
            
            batch: List[Tuple[str, Optional[str], Optional[str], Optional[str], str, int]] = []
            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                if loading_cancelled:
                    con.rollback()
                    raise RuntimeError("loading cancelled")
                    
                word_raw = row[word_col_idx] if word_col_idx < len(row or ()) else None
                display_word = "" if word_raw is None else str(word_raw).strip()
                norm = normalize_word(display_word)
                
                phonetic_val = row[2] if (row and len(row) > 2) else None
                meaning_val = row[3] if (row and len(row) > 3) else None
                phonetic = None if phonetic_val is None else str(phonetic_val)
                meaning = None if meaning_val is None else str(meaning_val)

                batch.append((norm, display_word or None, phonetic, meaning, sheet, int(row_idx)))

                processed_now = loading_state.increment_processed(
                    increment=1,
                    sample_word=display_word,
                    sample_step=SAMPLE_STEP,
                    latest_limit=LATEST_LIMIT,
                )

                if len(batch) >= 10000:
                    cur.executemany(insert_sql, batch)
                    batch.clear()

            if batch:
                cur.executemany(insert_sql, batch)
                batch.clear()

        cur.execute("CREATE INDEX idx_entries_word_norm ON entries(word_norm);")
        cur.execute("CREATE INDEX idx_entries_sheet_row ON entries(sheet, row_index);")

        con.commit()
        
        # 加载完成后，关闭WAL模式并合并临时文件
        logger.info("正在合并WAL文件到主数据库...")
        cur.execute("PRAGMA wal_checkpoint(TRUNCATE);")  # 强制合并WAL到主文件
        cur.execute("PRAGMA journal_mode=DELETE;")      # 切换回标准模式
        con.commit()
        
        # 优化数据库文件（可选，会花一点时间）
        logger.info("正在优化数据库文件...")
        cur.execute("VACUUM;")
        logger.info("数据库优化完成，WAL模式已关闭")

        app_state.data_loaded = True
        app_state.current_excel_file = file_path
        logger.info(f"Excel数据加载完成: {os.path.basename(file_path)}")
    finally:
        con.close()


def loader_worker(file_path: str):
    """后台加载线程"""
    global loading_thread, loading_cancelled
    from .app_state import app_state
    
    try:
        file_name = os.path.basename(file_path)
        total_rows = compute_total_rows(file_path)
        loading_state.reset_for_file(file_name, total_rows)
        loading_cancelled = False
        logger.info(f"开始加载Excel: {file_name}, 总行数: {total_rows}")
        _rebuild_sqlite_from_excel(file_path)
    except Exception as exc:
        logger.error(f"Excel加载失败: {exc}", exc_info=True)
        loading_state.mark_finished(error=str(exc))
        app_state.data_loaded = False
    finally:
        loading_state.mark_finished()
        loading_thread = None


def is_data_loaded() -> bool:
    """检查数据是否已加载"""
    from .app_state import app_state
    
    if not app_state.data_loaded:
        # 尝试检测数据库是否存在
        try:
            if os.path.exists(SQLITE_DB_PATH):
                # 使用连接池检查（如果已初始化）
                if app_state.db_pool:
                    try:
                        with app_state.db_pool.get_db() as conn:
                            cur = conn.cursor()
                            cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='entries'")
                            exists = cur.fetchone() is not None
                            if exists:
                                app_state.data_loaded = True
                    except Exception:
                        app_state.data_loaded = False
                else:
                    # 连接池未初始化，直接连接检查
                    con = sqlite3.connect(SQLITE_DB_PATH)
                    cur = con.cursor()
                    cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='entries'")
                    exists = cur.fetchone() is not None
                    con.close()
                    if exists:
                        app_state.data_loaded = True
        except Exception:
            app_state.data_loaded = False
    return app_state.data_loaded


def start_loading(file_path: str) -> bool:
    """启动加载线程"""
    global loading_thread
    from .app_state import app_state
    
    if loading_state.snapshot().get("running"):
        logger.warning("已有加载任务在运行")
        return False
    
    app_state.data_loaded = False
    app_state.current_excel_file = None
    
    # 不删除数据库文件，直接用 DROP TABLE 覆盖表内容（避免文件占用问题）
    logger.info("准备加载新数据")
        
    t = threading.Thread(target=loader_worker, args=(file_path,), daemon=True)
    loading_thread = t
    t.start()
    logger.info(f"启动加载线程: {os.path.basename(file_path)}")
    return True


def unload_data():
    """卸载数据"""
    from .app_state import app_state
    
    app_state.data_loaded = False
    app_state.current_excel_file = None
    
    try:
        if os.path.exists(SQLITE_DB_PATH):
            os.remove(SQLITE_DB_PATH)
            logger.info("数据已卸载")
    except Exception as e:
        logger.warning(f"卸载数据失败: {e}")

